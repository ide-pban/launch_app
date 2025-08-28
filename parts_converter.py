import streamlit as st
import pandas as pd
import openpyxl
import PyPDF2
import re
import io
import os
from typing import Dict, List, Tuple, Optional
import json
import shutil
from component_classifier import ComponentClassifier

class PartsListConverter:
    def __init__(self):
        self.template_path = 'parts_list_template.xlsx'
        self.main_sheet_name = '部品リスト入力フォーム'  # Main parts list input form sheet
        self.classifier = ComponentClassifier()
        
    def extract_from_excel(self, uploaded_file) -> pd.DataFrame:
        """Extract data from Excel file"""
        try:
            # Try to read all sheets
            excel_data = pd.read_excel(uploaded_file, sheet_name=None, header=None)
            
            # Combine data from all sheets
            all_data = []
            for sheet_name, df in excel_data.items():
                if not df.empty:
                    # Add sheet name as a column for reference
                    df['source_sheet'] = sheet_name
                    all_data.append(df)
            
            if all_data:
                combined_df = pd.concat(all_data, ignore_index=True)
                return self.normalize_dataframe(combined_df)
            else:
                return pd.DataFrame()
                
        except Exception as e:
            st.error(f"Excelファイルの読み込みエラー: {str(e)}")
            return pd.DataFrame()
    
    def extract_from_csv(self, uploaded_file) -> pd.DataFrame:
        """Extract data from CSV file"""
        try:
            # Try different encodings
            encodings = ['utf-8', 'shift-jis', 'cp932', 'iso-2022-jp']
            
            for encoding in encodings:
                try:
                    uploaded_file.seek(0)
                    df = pd.read_csv(uploaded_file, encoding=encoding, header=None)
                    return self.normalize_dataframe(df)
                except UnicodeDecodeError:
                    continue
                except Exception:
                    continue
            
            st.error("CSVファイルの文字コードを自動判定できませんでした")
            return pd.DataFrame()
            
        except Exception as e:
            st.error(f"CSVファイルの読み込みエラー: {str(e)}")
            return pd.DataFrame()
    
    def extract_from_pdf(self, uploaded_file) -> pd.DataFrame:
        """Extract data from PDF file (basic text extraction)"""
        try:
            reader = PyPDF2.PdfReader(uploaded_file)
            all_text = ""
            
            for page in reader.pages:
                all_text += page.extract_text() + "\n"
            
            # Split text into lines and try to parse as tabular data
            lines = [line.strip() for line in all_text.split('\n') if line.strip()]
            
            # Simple heuristic: assume each line is a record
            data = []
            for line in lines:
                # Split by common delimiters
                parts = re.split(r'[\s\t,;]+', line)
                data.append(parts)
            
            if data:
                df = pd.DataFrame(data)
                return self.normalize_dataframe(df)
            else:
                return pd.DataFrame()
                
        except Exception as e:
            st.error(f"PDFファイルの読み込みエラー: {str(e)}")
            return pd.DataFrame()
    
    def normalize_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Normalize dataframe for consistent processing"""
        if df.empty:
            return df
        
        # Convert all columns to string and fill NaN with empty string
        df = df.astype(str).fillna('')
        
        # Remove completely empty rows
        df = df.loc[~(df == '').all(axis=1)]
        
        return df
    
    def detect_part_numbers(self, df: pd.DataFrame) -> List[Tuple[int, str]]:
        """Detect manufacturer part numbers using pattern matching"""
        part_numbers = []
        
        # Enhanced part number patterns with priority
        patterns = [
            (r'[A-Z]{2,6}\d{2,8}[A-Z]*\d*[A-Z]*-?[A-Z]*\d*', 20),  # Like RK73B2ATTD1002F, ATmega328P-AU
            (r'[A-Z]+\d+[A-Z]+\d+[A-Z]*', 15),                     # Letter-number-letter-number
            (r'[A-Z]{3,}\d{2,}[A-Z]*', 10),                        # 3+ letters, 2+ numbers
            (r'[A-Z]+\d+[A-Z]*-[A-Z]*\d*', 12),                    # With dash
            (r'[A-Z]+\d+', 5),                                      # Simple alphanumeric
        ]
        
        for idx, row in df.iterrows():
            found_part = None
            confidence = 0
            
            for col_idx, cell_value in enumerate(row):
                if pd.isna(cell_value) or str(cell_value).strip() == '':
                    continue
                
                cell_str = str(cell_value).strip()
                
                # Skip obvious non-part-number fields
                if cell_str.lower() in ['item', 'manufacturer', 'quantity', 'description', 'reference']:
                    continue
                
                # Check against patterns with priority
                for pattern, base_score in patterns:
                    matches = re.findall(pattern, cell_str, re.IGNORECASE)
                    for match in matches:
                        # Calculate confidence based on length and pattern complexity
                        current_confidence = base_score + len(match)
                        
                        # Bonus for specific part number characteristics
                        if re.search(r'[A-Z]{2,}.*\d{3,}.*[A-Z]', match):  # Letters-numbers-letters
                            current_confidence += 10
                        
                        if len(match) > 8:  # Longer part numbers are usually more reliable
                            current_confidence += 5
                            
                        if current_confidence > confidence:
                            found_part = match
                            confidence = current_confidence
            
            if found_part:
                part_numbers.append((idx, found_part))
        
        return part_numbers
    
    def extract_reference_designators(self, df: pd.DataFrame, part_idx: int) -> str:
        """Extract reference designators (like R1, R5, C10) from the row and format with commas"""
        row = df.iloc[part_idx]
        ref_patterns = [
            r'[RCLUDQJXYFT]\d+(?:[,;]\s*[RCLUDQJXYFT]\d+)*',  # R1,R5,C10 or R1;R5;C10 format
            r'[RCLUDQJXYFT]\d+(?:\s*[,;]\s*[RCLUDQJXYFT]\d+)*',  # With spaces
            r'[RCLUDQJXYFT]\d+(?:-[RCLUDQJXYFT]\d+)*',  # Range format like R1-R5
        ]
        
        best_match = ""
        max_refs = 0
        
        for cell_value in row:
            if pd.isna(cell_value):
                continue
            
            cell_str = str(cell_value).strip()
            for pattern in ref_patterns:
                matches = re.findall(pattern, cell_str, re.IGNORECASE)
                for match in matches:
                    # Count how many references this match contains
                    ref_count = len(re.findall(r'[RCLUDQJXYFT]\d+', match))
                    if ref_count > max_refs:
                        max_refs = ref_count
                        best_match = match
        
        # Format the references to always use commas as separators
        if best_match:
            # Extract individual references and join with commas
            individual_refs = re.findall(r'[RCLUDQJXYFT]\d+', best_match, re.IGNORECASE)
            return ','.join(individual_refs)
        
        return best_match
    
    def count_references(self, ref_string: str) -> int:
        """Count the number of reference designators"""
        if not ref_string:
            return 0
        
        # Split by comma and count
        refs = [r.strip() for r in ref_string.split(',') if r.strip()]
        return len(refs)
    
    def calculate_quantity(self, ref_count: int, required_qty: Optional[int], panel_count: int) -> int:
        """Calculate required quantity based on reference count and panel count"""
        if required_qty is not None and required_qty > 0:
            # Use provided quantity divided by panel count
            return max(1, int(required_qty / panel_count))
        elif ref_count > 0:
            # Use reference count divided by panel count
            return max(1, int(ref_count / panel_count))
        else:
            return 1  # Default to 1
    
    def detect_manufacturers(self, df: pd.DataFrame) -> Dict[int, str]:
        """Detect manufacturer names in the data"""
        manufacturers = {}
        common_manufacturers = [
            'KOA', 'Murata', 'TDK', 'Panasonic', 'Vishay', 'Yageo', 'Rohm', 'Taiyo Yuden',
            'Samsung', 'Nichicon', 'Rubycon', 'KEMET', 'AVX', 'Bourns', 'Coilcraft'
        ]
        
        for idx, row in df.iterrows():
            for cell_value in row:
                if pd.isna(cell_value):
                    continue
                
                cell_str = str(cell_value).strip()
                for manufacturer in common_manufacturers:
                    if manufacturer.lower() in cell_str.lower():
                        manufacturers[idx] = manufacturer
                        break
        
        return manufacturers
    
    def map_to_template(self, df: pd.DataFrame, panel_count: int) -> pd.DataFrame:
        """Map extracted data to template format"""
        # Detect part numbers and other data
        part_numbers = self.detect_part_numbers(df)
        manufacturers = self.detect_manufacturers(df)
        
        # Create template data
        template_data = []
        
        for idx, (row_idx, part_number) in enumerate(part_numbers, 1):
            ref_designators = self.extract_reference_designators(df, row_idx)
            ref_count = self.count_references(ref_designators)
            manufacturer = manufacturers.get(row_idx, '')
            
            # Calculate quantities
            qty_1_unit = self.calculate_quantity(ref_count, None, panel_count)
            qty_total = qty_1_unit * panel_count
            
            template_row = {
                'No': idx,
                'メーカー': manufacturer,
                '品名': '',  # Will be filled with defaults  
                '電子部品型番': part_number,
                '配置記号': ref_designators,
                '個': qty_total,
                '実装数': qty_1_unit,
                '合計': qty_total,
                '実装/検査': '実装',
                '部品型番': 'SMD',  # Default to SMD
                '必要数': qty_1_unit
            }
            
            # Enhance with component classification and defaults
            template_row = self.classifier.enhance_component_data(template_row)
            
            template_data.append(template_row)
        
        return pd.DataFrame(template_data)
    
    def create_output_file(self, mapped_data: pd.DataFrame, output_filename: str, panel_count: int = 8) -> str:
        """Create output file by copying template and replacing only data rows"""
        try:
            # Copy the original template to preserve exact formatting
            output_path = output_filename
            shutil.copy2(self.template_path, output_path)
            
            # Load the copied workbook but WITHOUT VBA to avoid Excel compatibility issues
            workbook = openpyxl.load_workbook(output_path, keep_vba=False, data_only=False)
            
            # Get the main parts list sheet
            sheet = None
            for name in workbook.sheetnames:
                if '部品リスト' in name and '入力' in name:
                    sheet = workbook[name]
                    break
                    
            if not sheet:
                st.error(f"Main sheet not found. Available: {workbook.sheetnames}")
                return None
            
            # Preserve formatting for merged cells in M1:M3 and O1:O3 before clearing data
            m_cells_formatting = {}
            o_cells_formatting = {}
            try:
                # Save formatting for M column merged cells (rows 1-3)
                for row_num in [1, 2, 3]:
                    cell = sheet.cell(row_num, 13)  # M column
                    if cell.value or hasattr(cell, '_style'):
                        m_cells_formatting[row_num] = {
                            'font': cell.font,
                            'fill': cell.fill,
                            'border': cell.border,
                            'alignment': cell.alignment,
                            'number_format': cell.number_format
                        }
                
                # Save formatting for O column merged cells (rows 1-3)
                for row_num in [1, 2, 3]:
                    cell = sheet.cell(row_num, 15)  # O column
                    if cell.value or hasattr(cell, '_style'):
                        o_cells_formatting[row_num] = {
                            'font': cell.font,
                            'fill': cell.fill,
                            'border': cell.border,
                            'alignment': cell.alignment,
                            'number_format': cell.number_format
                        }
            except Exception as e:
                print(f"Error saving formatting: {e}")
            
            # Clear ONLY the data rows (7 and below) without touching formatting
            # This preserves all the template structure while clearing old data
            data_start_row = 7
            max_clear_row = data_start_row + 50  # Clear reasonable range
            
            for row_num in range(data_start_row, max_clear_row):
                for col_num in range(1, 25):  # Clear enough columns
                    try:
                        cell = sheet.cell(row_num, col_num)
                        # Only clear actual values, preserve formatting and formulas
                        if (cell.value is not None and 
                            not (isinstance(cell.value, str) and cell.value.startswith('='))):
                            cell.value = None
                    except Exception:
                        continue
            
            # Set panel count in M4
            try:
                sheet.cell(4, 13).value = panel_count  # M4: パネル枚数
            except Exception as e:
                print(f"Error setting panel count: {e}")
            
            # Restore preserved formatting for M column merged cells (rows 1-3)
            try:
                from openpyxl.styles import Font, Fill, Border, Alignment
                for row_num, formatting in m_cells_formatting.items():
                    cell = sheet.cell(row_num, 13)  # M column
                    if formatting:
                        cell.font = formatting['font']
                        cell.fill = formatting['fill']
                        cell.border = formatting['border']
                        cell.alignment = formatting['alignment']
                        cell.number_format = formatting['number_format']
            except Exception as e:
                print(f"Error restoring M column formatting: {e}")
            
            # Restore preserved formatting for O column merged cells (rows 1-3)
            try:
                for row_num, formatting in o_cells_formatting.items():
                    cell = sheet.cell(row_num, 15)  # O column
                    if formatting:
                        cell.font = formatting['font']
                        cell.fill = formatting['fill']
                        cell.border = formatting['border']
                        cell.alignment = formatting['alignment']
                        cell.number_format = formatting['number_format']
            except Exception as e:
                print(f"Error restoring O column formatting: {e}")
            
            # Preserve P板.com text color formatting in O1, O2, O3
            try:
                # Get current formatting for P板.com cells and preserve it
                for row_num in [1, 2, 3]:
                    cell = sheet.cell(row_num, 15)  # O column
                    if cell.value and 'P板.com' in str(cell.value):
                        # Preserve existing font color by not modifying the cell
                        pass
            except Exception as e:
                print(f"Error preserving P板.com formatting: {e}")
            
            # Preserve P column text color formatting in rows 1-4
            try:
                # Get current formatting for P column cells and preserve it
                for row_num in [1, 2, 3, 4]:
                    cell = sheet.cell(row_num, 16)  # P column
                    if cell.value:
                        # Preserve existing font color by not modifying the cell
                        pass
            except Exception as e:
                print(f"Error preserving P column formatting: {e}")
            
            # Insert new data starting from row 7
            for idx, row in mapped_data.iterrows():
                current_row = data_start_row + idx
                
                try:
                    # Check if cells are merged before setting values
                    def set_cell_value_safe(row_num, col_num, value):
                        try:
                            cell = sheet.cell(row_num, col_num)
                            # Skip merged cells (they are read-only)
                            if hasattr(cell, '__class__') and 'MergedCell' in str(cell.__class__):
                                return
                            cell.value = value
                        except Exception as e:
                            print(f"Warning: Could not set cell {row_num},{col_num}: {e}")
                    
                    # Set basic data - the template formatting will be preserved
                    set_cell_value_safe(current_row, 1, row.get('No', ''))           # A: No
                    set_cell_value_safe(current_row, 2, row.get('メーカー', ''))      # B: メーカー
                    set_cell_value_safe(current_row, 3, row.get('品名', ''))         # C: 部品種別  
                    set_cell_value_safe(current_row, 4, row.get('電子部品型番', ''))   # D: 電子部品型番
                    set_cell_value_safe(current_row, 5, row.get('配置記号', ''))      # E: 配置記号
                    
                    # F: Set quantity based on comma-separated references in E column
                    ref_string = row.get('配置記号', '')
                    if ref_string:
                        ref_count = len([r.strip() for r in ref_string.split(',') if r.strip()])
                    else:
                        ref_count = 0
                    set_cell_value_safe(current_row, 6, ref_count)  # F: 個数
                    set_cell_value_safe(current_row, 7, row.get('実装数', ''))        # G: 1個あたり
                    
                    # Column H: Total formula - use template's formula style
                    try:
                        h_cell = sheet.cell(current_row, 8)
                        if not (hasattr(h_cell, '__class__') and 'MergedCell' in str(h_cell.__class__)):
                            h_cell.value = f"=IF(F{current_row}=\"\",\"\",F{current_row}*G{current_row})"
                    except Exception:
                        pass
                    
                    set_cell_value_safe(current_row, 9, row.get('実装/検査', '実装'))   # I: 実装/未実装
                    
                    # SMD/DIP/Special classification - clear all first, then set appropriate one
                    set_cell_value_safe(current_row, 10, None)  # Clear SMD
                    set_cell_value_safe(current_row, 11, None)  # Clear DIP  
                    set_cell_value_safe(current_row, 12, None)  # Clear Special
                    
                    part_type = row.get('部品型番', 'SMD')
                    if part_type == 'SMD':
                        set_cell_value_safe(current_row, 10, 'SMD')
                    elif part_type == 'DIP':
                        set_cell_value_safe(current_row, 11, 'DIP')
                    else:
                        set_cell_value_safe(current_row, 12, '特殊（BGA等）')
                    
                    # Column M: 必要数
                    set_cell_value_safe(current_row, 13, row.get('必要数', row.get('実装数', '')))
                    
                except Exception as e:
                    print(f"Error setting data in row {current_row}: {e}")
                    continue
            
            # Save without VBA for Excel compatibility
            workbook.save(output_path)
            workbook.close()
            
            # Verify the file was created
            if not os.path.exists(output_path):
                raise Exception("Output file was not created successfully")
                
            return output_path
            
        except Exception as e:
            import traceback
            error_msg = f"Output file creation error: {str(e)}\n{traceback.format_exc()}"
            st.error(f"出力ファイル作成エラー: {str(e)}")
            print(error_msg)
            return None
    

def main():
    st.set_page_config(
        page_title="部品リスト変換アプリ",
        page_icon="📋",
        layout="wide"
    )
    
    st.title("📋 部品リスト変換アプリ")
    st.markdown("Excel、CSV、PDFファイルを部品リストフォーマットに自動変換します")
    
    # Initialize converter
    converter = PartsListConverter()
    
    # Sidebar for settings
    with st.sidebar:
        st.header("設定")
        
        panel_count = st.number_input(
            "パネル枚数",
            min_value=1,
            value=8,
            help="部品数量計算に使用するパネル枚数を指定してください"
        )
        
        
        st.header("使用方法")
        st.markdown("""
        1. ファイルをアップロードしてください（Excel、CSV、PDF対応）
        2. パネル枚数を設定してください
        3. 「変換実行」ボタンを押してください
        4. 変換結果をダウンロードしてください
        """)
    
    # File upload
    st.header("ファイルアップロード")
    uploaded_file = st.file_uploader(
        "変換するファイルを選択してください",
        type=['xlsx', 'xls', 'csv', 'pdf'],
        help="Excel、CSV、またはPDFファイルをアップロードしてください"
    )
    
    if uploaded_file is not None:
        st.success(f"ファイル '{uploaded_file.name}' がアップロードされました")
        
        # Extract data based on file type
        file_extension = uploaded_file.name.split('.')[-1].lower()
        
        if st.button("変換実行", type="primary"):
            with st.spinner("ファイルを処理中..."):
                
                if file_extension in ['xlsx', 'xls']:
                    extracted_data = converter.extract_from_excel(uploaded_file)
                elif file_extension == 'csv':
                    extracted_data = converter.extract_from_csv(uploaded_file)
                elif file_extension == 'pdf':
                    extracted_data = converter.extract_from_pdf(uploaded_file)
                else:
                    st.error("サポートされていないファイル形式です")
                    return
                
                if extracted_data.empty:
                    st.warning("ファイルからデータを抽出できませんでした")
                    return
                
                st.header("抽出されたデータ")
                st.dataframe(extracted_data.head(10))
                
                # Map to template format
                with st.spinner("テンプレート形式に変換中..."):
                    mapped_data = converter.map_to_template(extracted_data, panel_count)
                
                if not mapped_data.empty:
                    st.header("変換結果プレビュー")
                    st.dataframe(mapped_data)
                    
                    # Create Excel output file
                    output_filename = f"converted_{uploaded_file.name.split('.')[0]}.xlsx"
                    
                    with st.spinner("出力ファイルを作成中..."):
                        output_path = converter.create_output_file(mapped_data, output_filename, panel_count)
                    
                    if output_path and os.path.exists(output_path):
                        st.success("✅ 変換完了！")
                        
                        # Download button
                        with open(output_path, "rb") as file:
                            st.download_button(
                                label="📁 変換結果をダウンロード",
                                data=file.read(),
                                file_name=output_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        
                        # Show statistics
                        st.info(f"部品数: {len(mapped_data)}点 | パネル枚数: {panel_count}枚")
                        
                    else:
                        st.error("出力ファイルの作成に失敗しました")
                else:
                    st.warning("変換可能な部品データが見つかりませんでした")

if __name__ == "__main__":
    main()
